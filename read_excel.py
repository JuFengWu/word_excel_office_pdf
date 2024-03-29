import openpyxl
wb = openpyxl.load_workbook('excel_data\ex1.xlsx')     # 開啟 Excel 檔案

names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
s1 = wb['工作表1']        # 取得工作表名稱為「工作表1」的內容

print(names)
print(s1['A1'].value)        # 取出 A1 的內容
print(s1.cell(1, 1).value)   # 等同取出 A1 的內容
print(s1['B2'].value)        # 取出 B2 的內容
print(s1.cell(2, 2).value)   # 等同取出 B2 的內容


s1.cell(6,2).value = 400 #將資料寫入B6
wb.save("excel_data\ex1.xlsx")
wb.save("new.xlsx")
